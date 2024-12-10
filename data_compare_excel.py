import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Input data dictionary
recon_data = {
    'sybase_iqa_counts': {
        'orders': {
            '2024-01-02': 2343,
            '2024-01-03': 2345
        },
        'transactions': {
            '2024-01-02': 2343,
            '2024-01-03': 2345
        }
    },
    'sybase_iq_counts': {
        'orders': {
            ' 2024-01-02': 2343,
            ' 2024-01-03': 2345,
            ' 2024-12-01': 1223
        },
        'transactions': {
            '2024-01-02': 2343,
            '2024-01-05': 2345
        }
    }
}

# Helper function to convert dictionary to DataFrame
def dict_to_dataframe(data):
    rows = []
    for category, dates in data.items():
        for date, count in dates.items():
            rows.append({'Category': category, 'Date': date.strip(), 'Count': count})
    return pd.DataFrame(rows)

# Convert sections to DataFrames
iqa_df = dict_to_dataframe(recon_data['sybase_iqa_counts'])
iq_df = dict_to_dataframe(recon_data['sybase_iq_counts'])

# Perform an outer merge to find mismatches
merged_df = pd.merge(
    iqa_df,
    iq_df,
    on=['Category', 'Date'],
    how='outer',
    suffixes=('_iqa', '_iq'),
    indicator=True
)

# Handle NaN for category columns and replace with NULL
for col in merged_df.select_dtypes(include=['category']).columns:
    merged_df[col] = merged_df[col].cat.add_categories('NULL')  # Add 'NULL' to categories
merged_df.fillna('NULL', inplace=True)

# Add a column to indicate mismatches
merged_df['Mismatch'] = (merged_df['_merge'] != 'both') | (merged_df['Count_iqa'] != merged_df['Count_iq'])

# Select relevant columns for display
highlighted_df = merged_df[['Category', 'Date', 'Count_iqa', 'Count_iq', 'Mismatch']]

# Write to Excel with color coding
excel_file_path = "recon_diff.xlsx"
with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
    highlighted_df.to_excel(writer, index=False, sheet_name='Comparison')
    
    # Apply color formatting
    workbook = writer.book
    worksheet = writer.sheets['Comparison']
    
    # Define color fills
    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    
    # Iterate over the rows to apply styling
    for row_idx, (_, row) in enumerate(highlighted_df.iterrows(), start=2):  # Start from row 2 to skip headers
        fill = red_fill if row['Mismatch'] else green_fill
        for col_idx in range(1, len(row) + 1):  # Iterate over all columns
            worksheet.cell(row=row_idx, column=col_idx).fill = fill

print(f"Comparison results saved to {excel_file_path}")
